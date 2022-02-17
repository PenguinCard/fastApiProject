from fastapi import APIRouter, File, UploadFile, Response

from openpyxl import load_workbook, Workbook
from PyPDF2 import PdfFileReader, PdfFileWriter

from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import letter

from io import BytesIO
from bs4 import BeautifulSoup
from zipfile import ZipFile

import re
import math
import pdfplumber

from typing import List

# spark
from pyspark.sql import SparkSession
from pyspark.sql.types import StructType, StructField, StringType
from pyspark.sql import Row

router = APIRouter(prefix="/api/file", tags=["file"])

# add Font
pdfmetrics.registerFont(TTFont("D2Coding", "D2Coding.ttf"))

# 스파크 세션 생성
spark = SparkSession.builder.config('spark.driver.host', '127.0.0.1').getOrCreate()

# spark로 json 파일을 읽은 뒤 SQL View(이름은 'SKU')로 생성
spark.read.json('./config/SKU.json', multiLine=True).createOrReplaceTempView("SKU")

@router.get("/")
async def img_root():
    return "file root"

@router.post("/sparktopdf")
async def sparktopdf(files: List[UploadFile] = File(...)):
    
    xlsx_files = list(filter(lambda file: True if file.filename.find('.xlsx') >= 0 else False, files))
    html_files = list(filter(lambda file: True if file.filename.find('.html') >= 0 else False, files))
    pdf_files = list(filter(lambda file: True if file.filename.find('.pdf') >= 0 else False, files))
    
    zio = BytesIO()
    zips = ZipFile(zio, 'w')
    
    # 송장감지 텍스트
    regex = re.compile(r'LZDID|First Mile Warehouse')
    
    # 생성할 spark View Schema 정의
    Schema = StructType([
        StructField("orderNumber", StringType(), True),
        StructField("orderSku", StringType(), True)
    ])
    
    # spark 에 구성될 행 데이터
    rows = []
    
    # xlsx 파일 갯수만큼 반복
    for xlsx_file in xlsx_files:
        idx = 2
        
        wb = load_workbook(filename=BytesIO(xlsx_file.file.read()))
        ws = wb.active
        
        # M열과 F열의 데이터가 둘 중에 하나라도 없으면 break
        while ws["M{}".format(idx)].value is not None and ws["F{}".format(idx)].value is not None:
            # 주문번호
            order_number = str(ws["M{}".format(idx)].value)
            # text 정렬작업
            text = str(ws["F{}".format(idx)].value)
            text = re.sub(r'\)\s1', ')', text)
            text = re.sub(r'\(\d{4}\)|[+]', ',', text)
            text = text.strip()
            texts = text.split(',')
            texts = list(filter(None, texts))
            texts = list(map(lambda s: s.strip(), texts))

            # spark View에 행 추가
            for text in texts:
                rows.append(
                    Row(
                        order_number,
                        text
                    )
                )

            idx = idx + 1
    
    spark.createDataFrame(rows, Schema).createOrReplaceTempView("ORDERSKU")
    
    for pdf_file in pdf_files:
        baseName = re.sub(r"\.pdf", "", pdf_file.filename)
        pdf_file = list(filter(lambda file: True if file.filename.find('{}.pdf'.format(baseName)) >= 0 else False, pdf_files))
        pdf_file = pdf_file[0]
        html_file = list(filter(lambda file: True if file.filename.find('{}.html'.format(baseName)) >= 0 else False, html_files))
        html_file = html_file[0]
        html = html_file.file.read()
        dom = BeautifulSoup(html, 'html.parser')
        
        # html 파일에서 order_numbers 담는 리스트
        order_numbers = []

        # order_number 이 있는 셀렉터
        contents = dom.select('table tbody tr td p')
        # 셀렉터 반복하여 order_number 추출
        for content in contents:
            text = re.findall(r'Order Number[:] \d+', content.text)
            if len(text) > 0:
                order_number = re.sub(r'Order Number[:]', '', text[0])
                order_numbers.append(order_number.strip())

        
        pdf_data = BytesIO(pdf_file.file.read())
        
        pdf = PdfFileReader(pdf_data)
        pages = pdfplumber.open(pdf_data).pages
        # pdf 작성
        output = PdfFileWriter()
        
        idx = 0
        
        for pageNum, page in enumerate(pages):
            # 텍스트를 추출하여 송장 감지
            texts = page.extract_text()
            check_invoice = regex.search(texts)

            if check_invoice is not None:
                query_order = "SELECT orderSku FROM ORDERSKU WHERE orderNumber like '{}'".format(order_numbers[idx])
                df = spark.sql(query_order)
                products = {}
                for row in df.rdd.collect():
                    query_sku = "SELECT * FROM SKU WHERE NAME like '%{}%'".format(row.orderSku)
                    df2 = spark.sql(query_sku)
                    try:
                        print(df2.rdd.take(1))
                        row_data = df2.rdd.take(1)[0]
                        data_text = "{} ({})".format(row_data.NAME, row_data.CODE)
                        if data_text in products:
                            products[data_text] = products[data_text] + 1
                        else:
                            products[data_text] = 1
                    except:
                        continue

                packet = BytesIO()
                can = canvas.Canvas(packet, pagesize=letter)  # pdf 출력 letter 사이즈
                can.setFillColorRGB(255, 255, 255)  # 캔버스 색상 흰 색
                can.setLineWidth(0.75)  # 라인 너비
                can.rect(6.45, 525, 272.3, 70, fill=1)  # 사각형 생성
                can.setFillColorRGB(0, 0, 0)  # 캔버스 색상 검정색

                i = 0  # 딕셔너리 index
                for key, value in products.items():
                    pos_y = 587.5 - (11 * math.floor(i / 3))  # text y축
                    pos_x = 8  # text x축
                    if i % 3 == 0:
                        pos_x = 8
                    elif i % 3 == 1:
                        pos_x = 100
                    else:
                        pos_x = 192

                    write_text = '{} {}'.format(key, value)
                    if len(write_text) < 15:
                        can.setFont("D2Coding", 8)  # 폰트종류: D2Coding, 폰트크기: 8
                    elif len(write_text) >= 15 & len(write_text) < 17:
                        can.setFont("D2Coding", 7)  # 폰트종류: D2Coding, 폰트크기: 7
                    elif len(write_text) >= 17 & len(write_text) < 19:
                        can.setFont("D2Coding", 6)  # 폰트종류: D2Coding, 폰트크기: 6
                    else:
                        can.setFont("D2Coding", 5)  # 폰트종류: D2Coding, 폰트크기: 5

                    can.drawString(pos_x, pos_y, write_text)
                    i = i + 1

                can.save()
                packet.seek(0)
                new_pdf = PdfFileReader(packet)
                invoice_page = pdf.getPage(pageNum)
                invoice_page.mergePage(new_pdf.getPage(0))
                output.addPage(invoice_page)

                idx = idx + 1
                
        bio = BytesIO()
        output.write(bio)
        
        zips.writestr('{}.pdf'.format(baseName), bio.getbuffer())
    
    zips.close()
    headers = { 'Content-Disposition': 'attachment; filename="union.zip"' }
        
    return Response(zio.getvalue(), headers=headers)

@router.post("/xlsxtopdf")
async def xlsxtopdf(files: List[UploadFile] = File(...)):
    
    # filtering text
    regex = re.compile(r'\d*[가-힣].*')
    regex_count = re.compile(r'數量: \d+')
    
    xlsx_files = list(filter(lambda file: True if file.filename.find('.xlsx') >= 0 else False, files))
    pdf_files = list(filter(lambda file: True if file.filename.find('.pdf') >= 0 else False, files))
    
    zio = BytesIO()
    zips = ZipFile(zio, 'w')
    
    for xlsx_file in xlsx_files:
        # Write pdf
        output = PdfFileWriter()
        
        file_name = re.sub(".xlsx", "", xlsx_file.filename)
        pdf_file = list(filter(lambda file: True if file.filename.find('{}.pdf'.format(file_name)) >= 0 else False, pdf_files))
        pdf_file = pdf_file[0]
        
        wb = load_workbook(filename=BytesIO(xlsx_file.file.read()))
        ws = wb.active
        pdf = PdfFileReader(BytesIO(pdf_file.file.read()))
        
        for num in range(pdf.getNumPages()):
            # col
            xlsx_num_data = "C{}".format(num + 2)
            str_list = str(ws[xlsx_num_data].value).split('\n')
            count_list = list(map(regex_count.findall, str_list))
            str_list = list(map(regex.findall, str_list))
            str_list = list(map(lambda s: s[0], str_list))
            # make tuple to lists
            str_list = list(zip(count_list, str_list))
            
            posY = 75
            posY_chk = posY

            # create pdf Text
            packet = BytesIO()
            can = canvas.Canvas(packet, pagesize=letter)
            can.setFont("D2Coding", 7)  # set Font: D2Coding, Size: 7
            
            # posY 미리 계산기능 추가
            for idx, (count, str_data) in enumerate(str_list):  # drawing Text, idx, str_data, exception case: G01305644544
                if posY_chk < 0:
                    posY_chk = posY_chk + 10
                    posY = posY + 10
                texts = re.sub(".*:|;.*", "", str_data).strip().split(',')
                text_len = math.ceil(len(texts) / 3)
                if text_len > 1:
                    for i in range(text_len):
                        posY_chk -= 10
                else:
                    posY_chk -= 10

            for idx, (count, str_data) in enumerate(str_list):  # drawing Text, idx, str_data, exception case: G01305644544
                texts = re.sub(".*:|;.*", "", str_data).strip().split(',')
                counts = re.sub('數量: ', " ", count[0])
                text_len = math.ceil(len(texts) / 3)
                if text_len > 1:
                    for i in range(text_len):
                        startPoint = i * 3
                        endPoint = (i + 1) * 3
                        write_text = ",".join(texts[startPoint:endPoint])
                        if i is text_len - 1:
                            write_text = write_text + counts
                        can.drawString(65, posY, write_text)
                        posY -= 10
                else:
                    can.drawString(65, posY, ",".join(texts) + counts)
                    posY -= 10
            can.save()
            packet.seek(0)
            new_pdf = PdfFileReader(packet)
            
            # get Page
            page = pdf.getPage(num)
            # page Merging
            page.mergePage(new_pdf.getPage(0))
            output.addPage(page)
            
            
        bio = BytesIO()
        output.write(bio)
            
        zips.writestr('{}.pdf'.format(file_name), bio.getbuffer())
        
    zips.close()
    headers = { 'Content-Disposition': 'attachment; filename="union.zip"' }
        
    return Response(zio.getvalue(), headers=headers)
    
    

# file 작업 multifile 파일 수신
@router.post("/merge_xlsx")
async def file_work(files: List[UploadFile] = File(...)):
    
    write_wb = Workbook()
    write_ws = write_wb.active
    
    xlsx_files = list(filter(lambda file: True if file.filename.find('.xlsx') >= 0 else False, files))
    
    for xlsx_file in xlsx_files:
        wb = load_workbook(filename=BytesIO(xlsx_file.file.read()))
        ws = wb.active
        
        idx = 1
        
        for row in ws.iter_rows(values_only=True):
            if idx != 1:
                write_ws.append(row)
            idx = idx + 1
            
    bio = BytesIO()
    write_wb.save(bio)
    
    headers = { 'Content-Disposition': 'attachment; filename="union.xlsx"' }
    
    return Response(bio.getvalue(), headers=headers)